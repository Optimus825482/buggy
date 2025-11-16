"""
Buggy Call - Reports Routes
Powered by Erkan ERDEM
"""
from flask import Blueprint, jsonify, request, session
from functools import wraps
from datetime import datetime, timedelta
from app import csrf
from app.models.user import SystemUser
from app.services.report_service import ReportService
from app.models.request import RequestStatus

reports_bp = Blueprint('reports', __name__)


def api_login_required(fn):
    """Decorator for API authentication"""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized', 'message': 'Lütfen giriş yapın'}), 401
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    """Decorator to ensure user is admin"""
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401

        user = SystemUser.query.get(session['user_id'])
        if not user or user.role != 'admin':
            return jsonify({'error': 'Forbidden', 'message': 'Admin yetkisi gerekli'}), 403

        return fn(*args, **kwargs)
    return wrapper


# ==================== Daily Summary ====================
@reports_bp.route('/daily-summary', methods=['GET'])
@api_login_required
def get_daily_summary():
    """Get daily summary report"""
    try:
        user = SystemUser.query.get(session['user_id'])

        # Get date from query params (defaults to today)
        date_str = request.args.get('date')
        if date_str:
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        else:
            date = datetime.utcnow()

        report = ReportService.get_daily_summary(user.hotel_id, date)

        return jsonify({
            'success': True,
            'report': report
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Buggy Performance ====================
@reports_bp.route('/buggy-performance', methods=['GET'])
@api_login_required
def get_buggy_performance():
    """Get buggy performance report"""
    try:
        user = SystemUser.query.get(session['user_id'])

        # Get optional parameters
        buggy_id = request.args.get('buggy_id', type=int)

        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        start_date = None
        end_date = None

        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400

        report = ReportService.get_buggy_performance(
            user.hotel_id,
            buggy_id,
            start_date,
            end_date
        )
      

        return jsonify({
            'success': True,
            'report': report
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Location Analytics ====================
@reports_bp.route('/location-analytics', methods=['GET'])
@api_login_required
def get_location_analytics():
    """Get location analytics report"""
    try:
        user = SystemUser.query.get(session['user_id'])

        # Get optional date range
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        start_date = None
        end_date = None

        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400

        report = ReportService.get_location_analytics(
            user.hotel_id,
            start_date,
            end_date
        )

        return jsonify({
            'success': True,
            'report': report
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Request Details ====================
@reports_bp.route('/request-details', methods=['GET'])
@api_login_required
def get_request_details():
    """Get detailed request list"""
    try:
        user = SystemUser.query.get(session['user_id'])

        # Get optional parameters
        status_str = request.args.get('status')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        limit = request.args.get('limit', 100, type=int)

        # Parse status
        status = None
        if status_str:
            try:
                status = RequestStatus(status_str)
            except ValueError:
                return jsonify({'error': f'Invalid status. Must be one of: {[s.value for s in RequestStatus]}'}), 400

        # Parse dates
        start_date = None
        end_date = None

        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400

        report = ReportService.get_request_details(
            user.hotel_id,
            status,
            start_date,
            end_date,
            limit
        )

        return jsonify({
            'success': True,
            'total': len(report),
            'requests': report
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Dashboard Stats ====================
@reports_bp.route('/dashboard-stats', methods=['GET'])
@api_login_required
def get_dashboard_stats():
    """Get real-time dashboard statistics"""
    try:
        from app.models.buggy import Buggy, BuggyStatus
        from app.models.request import BuggyRequest, RequestStatus
        from sqlalchemy import func

        user = SystemUser.query.get(session['user_id'])
        hotel_id = user.hotel_id

        # Active buggies count
        active_buggies = Buggy.query.filter_by(
            hotel_id=hotel_id,
            status=BuggyStatus.AVAILABLE
        ).count()

        # Pending requests count
        PENDING_requests = BuggyRequest.query.filter_by(
            hotel_id=hotel_id,
            status=RequestStatus.PENDING
        ).count()

        # Today's completed requests
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        today_completed = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.status == RequestStatus.COMPLETED,
            BuggyRequest.completed_at >= today_start
        ).count()

        # Average response time today
        today_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.status == RequestStatus.COMPLETED,
            BuggyRequest.requested_at >= today_start,
            BuggyRequest.accepted_at.isnot(None)
        ).all()

        avg_response_time = 0
        if today_requests:
            total_response = sum([
                (req.accepted_at - req.requested_at).total_seconds()
                for req in today_requests
            ])
            avg_response_time = round(total_response / len(today_requests) / 60, 2)  # in minutes

        return jsonify({
            'success': True,
            'stats': {
                'active_buggies': active_buggies,
                'PENDING_requests': PENDING_requests,
                'today_completed': today_completed,
                'avg_response_time_minutes': avg_response_time
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# ==================== Export to Excel ====================
@reports_bp.route('/export/excel/<report_type>', methods=['GET'])
@api_login_required
def export_excel(report_type):
    """Export report to Excel"""
    try:
        from flask import send_file
        from io import BytesIO
        
        user = SystemUser.query.get(session['user_id'])
        
        # Get report data based on type
        if report_type == 'daily-summary':
            date_str = request.args.get('date')
            date = datetime.strptime(date_str, '%Y-%m-%d') if date_str else datetime.utcnow()
            report_data = ReportService.get_daily_summary(user.hotel_id, date)
            data = [report_data]  # Convert dict to list
            filename = f'daily_summary_{date.strftime("%Y%m%d")}.xlsx'
            
        elif report_type == 'buggy-performance':
            buggy_id = request.args.get('buggy_id', type=int)
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d') if request.args.get('start_date') else None
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d') if request.args.get('end_date') else None
            data = ReportService.get_buggy_performance(user.hotel_id, buggy_id, start_date, end_date)
            filename = 'buggy_performance.xlsx'
            
        elif report_type == 'location-analytics':
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d') if request.args.get('start_date') else None
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d') if request.args.get('end_date') else None
            data = ReportService.get_location_analytics(user.hotel_id, start_date, end_date)
            filename = 'location_analytics.xlsx'
            
        elif report_type == 'request-details':
            status_str = request.args.get('status')
            status = RequestStatus(status_str) if status_str else None
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d') if request.args.get('start_date') else None
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d') if request.args.get('end_date') else None
            limit = request.args.get('limit', 1000, type=int)
            data = ReportService.get_request_details(user.hotel_id, status, start_date, end_date, limit)
            filename = 'request_details.xlsx'
            
        else:
            return jsonify({'error': 'Invalid report type'}), 400
        
        # Export to Excel
        excel_bytes = ReportService.export_to_excel(data, filename, sheet_name=report_type.replace('-', ' ').title())
        
        # Log export operation
        from app.services.audit_service import AuditService
        AuditService.log_action(
            action='report_exported',
            entity_type='report',
            new_values={
                'report_type': report_type,
                'format': 'excel',
                'filename': filename,
                'record_count': len(data)
            },
            user_id=session.get('user_id'),
            hotel_id=user.hotel_id
        )
        
        return send_file(
            BytesIO(excel_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Export to PDF ====================
@reports_bp.route('/export/pdf/<report_type>', methods=['GET'])
@api_login_required
def export_pdf(report_type):
    """Export report to PDF"""
    try:
        from flask import send_file
        from io import BytesIO
        
        user = SystemUser.query.get(session['user_id'])
        
        # Get report data based on type
        if report_type == 'daily-summary':
            date_str = request.args.get('date')
            date = datetime.strptime(date_str, '%Y-%m-%d') if date_str else datetime.utcnow()
            report_data = ReportService.get_daily_summary(user.hotel_id, date)
            data = [report_data]
            title = f'Daily Summary - {date.strftime("%Y-%m-%d")}'
            filename = f'daily_summary_{date.strftime("%Y%m%d")}.pdf'
            
        elif report_type == 'buggy-performance':
            buggy_id = request.args.get('buggy_id', type=int)
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d') if request.args.get('start_date') else None
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d') if request.args.get('end_date') else None
            data = ReportService.get_buggy_performance(user.hotel_id, buggy_id, start_date, end_date)
            title = 'Buggy Performance Report'
            filename = 'buggy_performance.pdf'
            
        elif report_type == 'location-analytics':
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d') if request.args.get('start_date') else None
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d') if request.args.get('end_date') else None
            data = ReportService.get_location_analytics(user.hotel_id, start_date, end_date)
            title = 'Location Analytics Report'
            filename = 'location_analytics.pdf'
            
        elif report_type == 'request-details':
            status_str = request.args.get('status')
            status = RequestStatus(status_str) if status_str else None
            start_date = datetime.strptime(request.args.get('start_date'), '%Y-%m-%d') if request.args.get('start_date') else None
            end_date = datetime.strptime(request.args.get('end_date'), '%Y-%m-%d') if request.args.get('end_date') else None
            limit = request.args.get('limit', 1000, type=int)
            data = ReportService.get_request_details(user.hotel_id, status, start_date, end_date, limit)
            title = 'Request Details Report'
            filename = 'request_details.pdf'
            
        else:
            return jsonify({'error': 'Invalid report type'}), 400
        
        # Export to PDF
        pdf_bytes = ReportService.export_to_pdf(data, title, filename)
        
        # Log export operation
        from app.services.audit_service import AuditService
        AuditService.log_action(
            action='report_exported',
            entity_type='report',
            new_values={
                'report_type': report_type,
                'format': 'pdf',
                'filename': filename,
                'record_count': len(data)
            },
            user_id=session.get('user_id'),
            hotel_id=user.hotel_id
        )
        
        return send_file(
            BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# ==================== Advanced Analytics ====================
@reports_bp.route('/advanced-analytics', methods=['GET'])
@api_login_required
def get_advanced_analytics():
    """
    Get advanced analytics including:
    - Unanswered requests (timeout after 1 hour)
    - Cancelled requests breakdown
    - Completed requests analysis
    - Response time trends
    - Peak hours analysis
    """
    try:
        user = SystemUser.query.get(session['user_id'])
        
        # Get date range from query params
        days = request.args.get('days', 30, type=int)
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            end_date = datetime.utcnow()
            start_date = end_date - timedelta(days=days)
        
        # Get comprehensive analytics
        analytics = ReportService.get_advanced_analytics(
            hotel_id=user.hotel_id,
            start_date=start_date,
            end_date=end_date
        )
        
        return jsonify({
            'success': True,
            'analytics': analytics,
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'days': (end_date - start_date).days
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== Timeout Statistics ====================
@reports_bp.route('/timeout-statistics', methods=['GET'])
@api_login_required
def get_timeout_statistics():
    """Get detailed statistics about timed out (unanswered) requests"""
    try:
        from app.tasks.timeout_checker import get_timeout_statistics
        
        user = SystemUser.query.get(session['user_id'])
        days = request.args.get('days', 30, type=int)
        
        stats = get_timeout_statistics(hotel_id=user.hotel_id, days=days)
        
        return jsonify({
            'success': True,
            'statistics': stats
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# ==================== Route Analytics ====================
@reports_bp.route('/route-analytics', methods=['GET'])
@api_login_required
def get_route_analytics():
    """
    Rota analizi - Başlangıç ve bitiş konumları arası detaylı istatistikler
    """
    try:
        user = SystemUser.query.get(session['user_id'])
        
        # Tarih aralığını al
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        # Varsayılan: Son 7 gün
        if not start_date_str:
            start_date = datetime.utcnow() - timedelta(days=7)
        else:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Geçersiz start_date formatı. YYYY-MM-DD kullanın'}), 400
        
        if not end_date_str:
            end_date = datetime.utcnow()
        else:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                # Günün sonuna kadar dahil et
                end_date = end_date + timedelta(days=1)
            except ValueError:
                return jsonify({'error': 'Geçersiz end_date formatı. YYYY-MM-DD kullanın'}), 400
        
        # Rota analizini al
        analytics = ReportService.get_route_analytics(user.hotel_id, start_date, end_date)
        
        return jsonify({
            'success': True,
            'data': analytics
        }), 200
        
    except Exception as e:
        import traceback
        print(f"❌ Route analytics error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ==================== Excel Export (Comprehensive) ====================
@reports_bp.route('/export-excel', methods=['POST'])
@csrf.exempt
@api_login_required
def export_comprehensive_excel():
    """
    Kapsamlı Excel raporu oluştur - Tüm istatistikler ve detaylar
    """
    try:
        from flask import send_file
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        user = SystemUser.query.get(session['user_id'])
        data = request.get_json()
        
        export_data = data.get('data', {})
        date_range = data.get('date_range', 'week')
        
        # Excel workbook oluştur
        wb = Workbook()
        
        # Stil tanımlamaları
        header_fill = PatternFill(start_color="1BA5A8", end_color="1BA5A8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="1A2B4A")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. ÖZET SAYFA
        ws_summary = wb.active
        ws_summary.title = "Özet"
        
        ws_summary['A1'] = "SHUTTLE CALL RAPORU"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1A2B4A")
        ws_summary.merge_cells('A1:D1')
        
        summary = export_data.get('summary', {})
        ws_summary['A3'] = "Tarih Aralığı:"
        ws_summary['B3'] = summary.get('date_range', '')
        ws_summary['A3'].font = Font(bold=True)
        
        # İstatistikler
        stats_data = [
            ['Metrik', 'Değer'],
            ['Toplam Talep', summary.get('total_requests', 0)],
            ['Tamamlanan', summary.get('completed', 0)],
            ['İptal Edilen', summary.get('cancelled', 0)],
            ['Bekleyen', summary.get('pending', 0)],
            ['Cevapsız', summary.get('unanswered', 0)],
            ['Başarı Oranı', summary.get('success_rate', '0%')],
            ['Ort. Yanıt Süresi', summary.get('avg_response_time', '0 dk')],
            ['Ort. Tamamlanma Süresi', summary.get('avg_completion_time', '0 dk')]
        ]
        
        for row_idx, row_data in enumerate(stats_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 5:  # Header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
        
        # Sütun genişlikleri
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # 2. DETAYLI TALEPLER
        ws_requests = wb.create_sheet("Talepler")
        
        headers = ['Tarih', 'Başlangıç', 'Bitiş', 'Shuttle', 'Sürücü', 'Oda', 'Yanıt (dk)', 'Tamamlanma (dk)', 'Durum']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_requests.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        requests = export_data.get('requests', [])
        for row_idx, req in enumerate(requests, start=2):
            ws_requests.cell(row=row_idx, column=1, value=req.get('tarih', '')).border = border
            ws_requests.cell(row=row_idx, column=2, value=req.get('baslangic', '')).border = border
            ws_requests.cell(row=row_idx, column=3, value=req.get('bitis', '')).border = border
            ws_requests.cell(row=row_idx, column=4, value=req.get('shuttle', '')).border = border
            ws_requests.cell(row=row_idx, column=5, value=req.get('surucu', '')).border = border
            ws_requests.cell(row=row_idx, column=6, value=req.get('oda', '')).border = border
            ws_requests.cell(row=row_idx, column=7, value=req.get('yanit_suresi_dk', 0)).border = border
            ws_requests.cell(row=row_idx, column=8, value=req.get('tamamlanma_suresi_dk', 0)).border = border
            ws_requests.cell(row=row_idx, column=9, value=req.get('durum', '')).border = border
        
        # Sütun genişlikleri
        for col in range(1, 10):
            ws_requests.column_dimensions[get_column_letter(col)].width = 18
        
        # 3. ROTA ANALİZİ
        route_analytics = export_data.get('route_analytics', {})
        if route_analytics and route_analytics.get('most_popular_routes'):
            ws_routes = wb.create_sheet("Rota Analizi")
            
            ws_routes['A1'] = "EN POPÜLER ROTALAR"
            ws_routes['A1'].font = title_font
            ws_routes.merge_cells('A1:E1')
            
            route_headers = ['Rota', 'Kullanım Sayısı', 'Ort. Süre (dk)', 'Min Süre (sn)', 'Max Süre (sn)']
            for col_idx, header in enumerate(route_headers, start=1):
                cell = ws_routes.cell(row=3, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            routes = route_analytics.get('most_popular_routes', [])
            for row_idx, route in enumerate(routes, start=4):
                ws_routes.cell(row=row_idx, column=1, value=route.get('route', '')).border = border
                ws_routes.cell(row=row_idx, column=2, value=route.get('count', 0)).border = border
                ws_routes.cell(row=row_idx, column=3, value=route.get('avg_time_minutes', 0)).border = border
                ws_routes.cell(row=row_idx, column=4, value=route.get('min_time_seconds', 0)).border = border
                ws_routes.cell(row=row_idx, column=5, value=route.get('max_time_seconds', 0)).border = border
            
            for col in range(1, 6):
                ws_routes.column_dimensions[get_column_letter(col)].width = 25
        
        # 4. SÜRÜCÜ PERFORMANSI
        if route_analytics and route_analytics.get('driver_performance'):
            ws_drivers = wb.create_sheet("Sürücü Performansı")
            
            ws_drivers['A1'] = "SÜRÜCÜ PERFORMANS RAPORU"
            ws_drivers['A1'].font = title_font
            ws_drivers.merge_cells('A1:D1')
            
            driver_headers = ['Sürücü', 'Tamamlanan', 'Ort. Süre (dk)', 'En Çok Kullanılan Rota']
            for col_idx, header in enumerate(driver_headers, start=1):
                cell = ws_drivers.cell(row=3, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            drivers = route_analytics.get('driver_performance', [])
            for row_idx, driver in enumerate(drivers, start=4):
                ws_drivers.cell(row=row_idx, column=1, value=driver.get('driver_name', '')).border = border
                ws_drivers.cell(row=row_idx, column=2, value=driver.get('total_completed', 0)).border = border
                ws_drivers.cell(row=row_idx, column=3, value=driver.get('avg_completion_time_minutes', 0)).border = border
                ws_drivers.cell(row=row_idx, column=4, value=driver.get('most_used_route', '')).border = border
            
            for col in range(1, 5):
                ws_drivers.column_dimensions[get_column_letter(col)].width = 30
        
        # 5. LOKASYON İSTATİSTİKLERİ
        location_stats = export_data.get('location_stats', {})
        if location_stats and location_stats.get('labels'):
            ws_locations = wb.create_sheet("Lokasyon İstatistikleri")
            
            ws_locations['A1'] = "LOKASYON BAZLI TALEP DAĞILIMI"
            ws_locations['A1'].font = title_font
            ws_locations.merge_cells('A1:B1')
            
            loc_headers = ['Lokasyon', 'Talep Sayısı']
            for col_idx, header in enumerate(loc_headers, start=1):
                cell = ws_locations.cell(row=3, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            labels = location_stats.get('labels', [])
            values = location_stats.get('values', [])
            for row_idx, (label, value) in enumerate(zip(labels, values), start=4):
                ws_locations.cell(row=row_idx, column=1, value=label).border = border
                ws_locations.cell(row=row_idx, column=2, value=value).border = border
            
            ws_locations.column_dimensions['A'].width = 30
            ws_locations.column_dimensions['B'].width = 15
        
        # Excel dosyasını kaydet
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Dosya adı
        filename = f'shuttle_call_rapor_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Audit log
        from app.services.audit_service import AuditService
        AuditService.log_action(
            action='excel_report_exported',
            entity_type='report',
            new_values={
                'date_range': date_range,
                'total_requests': summary.get('total_requests', 0),
                'filename': filename
            },
            user_id=session.get('user_id'),
            hotel_id=user.hotel_id
        )
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(f"❌ Excel export error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ==================== PDF Export (With Charts) ====================
@reports_bp.route('/export-pdf', methods=['POST'])
@csrf.exempt
@api_login_required
def export_comprehensive_pdf():
    """
    Kapsamlı PDF raporu oluştur - Grafikler ve tablolar dahil
    """
    try:
        from flask import send_file
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        import base64
        
        user = SystemUser.query.get(session['user_id'])
        data = request.get_json()
        
        export_data = data.get('data', {})
        date_range = data.get('date_range', 'week')
        
        # PDF buffer oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        # UTF-8 desteği için font kaydet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        
        # DejaVu Sans font'u kaydet (Türkçe karakter desteği)
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
        
        try:
            # Proje içindeki font yolu
            from flask import current_app
            font_dir = os.path.join(current_app.root_path, 'static', 'fonts')
            
            font_regular = os.path.join(font_dir, 'DejaVuSans.ttf')
            font_bold_file = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
            
            if os.path.exists(font_regular) and os.path.exists(font_bold_file):
                pdfmetrics.registerFont(TTFont('DejaVuSans', font_regular))
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_bold_file))
                font_name = 'DejaVuSans'
                font_bold = 'DejaVuSans-Bold'
                print(f"✅ DejaVu Sans fontları yüklendi: {font_dir}")
            else:
                print(f"⚠️ Font dosyaları bulunamadı: {font_dir}")
        except Exception as e:
            print(f"⚠️ Font yüklenemedi, Helvetica kullanılacak: {str(e)}")
            # Fallback: Helvetica (sınırlı Türkçe desteği)
        
        # Stiller
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1a2b4a'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=font_bold
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1BA5A8'),
            spaceAfter=12,
            spaceBefore=12,
            fontName=font_bold
        )
        
        # Normal stil için de font ayarla
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            fontName=font_name
        )
        
        # Başlık
        elements.append(Paragraph("SHUTTLE CALL RAPORU", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Tarih aralığı
        summary = export_data.get('summary', {})
        elements.append(Paragraph(f"<b>Tarih Aralığı:</b> {summary.get('date_range', '')}", normal_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Özet istatistikler tablosu
        elements.append(Paragraph("ÖZET İSTATİSTİKLER", heading_style))
        
        stats_data = [
            ['Metrik', 'Değer'],
            ['Toplam Talep', str(summary.get('total_requests', 0))],
            ['Tamamlanan', str(summary.get('completed', 0))],
            ['İptal Edilen', str(summary.get('cancelled', 0))],
            ['Bekleyen', str(summary.get('pending', 0))],
            ['Cevapsız', str(summary.get('unanswered', 0))],
            ['Başarı Oranı', summary.get('success_rate', '0%')],
            ['Ort. Yanıt Süresi', summary.get('avg_response_time', '0 dk')],
            ['Ort. Tamamlanma Süresi', summary.get('avg_completion_time', '0 dk')]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1BA5A8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        
        elements.append(stats_table)
        elements.append(PageBreak())
        
        # Grafikler
        charts = export_data.get('charts', {})
        
        if charts:
            elements.append(Paragraph("GRAFİKLER", heading_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Her grafik için
            chart_titles = {
                'daily': 'Günlük Trendler',
                'status': 'Durum Dağılımı',
                'location': 'Lokasyon Performansı',
                'hourly': 'Saatlik Yoğunluk',
                'route': 'Rota Analizi',
                'driver': 'Sürücü Performansı',
                'buggy': 'Shuttle Performansı'
            }
            
            for chart_key, chart_title in chart_titles.items():
                if chart_key in charts and charts[chart_key]:
                    try:
                        # Base64 image'i decode et
                        img_data = charts[chart_key].split(',')[1] if ',' in charts[chart_key] else charts[chart_key]
                        img_bytes = base64.b64decode(img_data)
                        img_buffer = BytesIO(img_bytes)
                        
                        # Grafik başlığı
                        elements.append(Paragraph(chart_title, heading_style))
                        
                        # Image ekle
                        img = Image(img_buffer, width=5*inch, height=3*inch)
                        elements.append(img)
                        elements.append(Spacer(1, 0.3*inch))
                        
                        # Her 2 grafikten sonra sayfa sonu
                        if chart_key in ['status', 'hourly', 'buggy']:
                            elements.append(PageBreak())
                    except Exception as e:
                        print(f"Grafik eklenemedi ({chart_key}): {str(e)}")
        
        # Son talepler tablosu
        elements.append(Paragraph("SON TALEPLER (İlk 50)", heading_style))
        elements.append(Spacer(1, 0.2*inch))
        
        requests = export_data.get('requests', [])[:50]
        if requests:
            table_data = [['Tarih', 'Başlangıç', 'Bitiş', 'Shuttle', 'Sürücü', 'Durum']]
            
            for req in requests:
                table_data.append([
                    req.get('tarih', '')[:16],  # Kısa tarih
                    req.get('baslangic', '')[:20],
                    req.get('bitis', '')[:20],
                    req.get('shuttle', ''),
                    req.get('surucu', '')[:15],
                    req.get('durum', '')
                ])
            
            requests_table = Table(table_data, colWidths=[1.2*inch, 1.3*inch, 1.3*inch, 0.8*inch, 1*inch, 0.9*inch])
            requests_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1BA5A8')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), font_bold),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
            ]))
            
            elements.append(requests_table)
        
        # PDF oluştur
        doc.build(elements)
        buffer.seek(0)
        
        # Dosya adı
        filename = f'shuttle_call_rapor_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        # Audit log
        from app.services.audit_service import AuditService
        AuditService.log_action(
            action='pdf_report_exported',
            entity_type='report',
            new_values={
                'date_range': date_range,
                'total_requests': summary.get('total_requests', 0),
                'filename': filename,
                'includes_charts': True
            },
            user_id=session.get('user_id'),
            hotel_id=user.hotel_id
        )
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(f"❌ PDF export error: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
