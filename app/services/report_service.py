"""
Reporting Service
Powered by Erkan ERDEM
"""
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_
from app import db
from app.models.request import BuggyRequest, RequestStatus
from app.models.buggy import Buggy, BuggyStatus
from app.models.location import Location
from typing import Dict, List, Optional, Any


class ReportService:
    """Service for generating reports and analytics"""

    @staticmethod
    def get_daily_summary(hotel_id: int, date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        Generate daily summary report

        Args:
            hotel_id: Hotel ID
            date: Date to report on (defaults to today)

        Returns:
            Dictionary with daily statistics
        """
        if date is None:
            date = datetime.utcnow()

        # Get start and end of day
        start_of_day = datetime(date.year, date.month, date.day, 0, 0, 0)
        end_of_day = start_of_day + timedelta(days=1)

        # Total requests
        total_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.requested_at >= start_of_day,
            BuggyRequest.requested_at < end_of_day
        ).count()

        # Completed requests
        completed_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.requested_at >= start_of_day,
            BuggyRequest.requested_at < end_of_day,
            BuggyRequest.status == RequestStatus.COMPLETED
        ).count()

        # Cancelled requests
        cancelled_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.requested_at >= start_of_day,
            BuggyRequest.requested_at < end_of_day,
            BuggyRequest.status == RequestStatus.CANCELLED
        ).count()

        # Pending requests
        PENDING_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.requested_at >= start_of_day,
            BuggyRequest.requested_at < end_of_day,
            BuggyRequest.status == RequestStatus.PENDING
        ).count()

        # Average response time (time to accept)
        completed_with_times = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.requested_at >= start_of_day,
            BuggyRequest.requested_at < end_of_day,
            BuggyRequest.status == RequestStatus.COMPLETED,
            BuggyRequest.accepted_at.isnot(None)
        ).all()

        avg_response_time = 0
        if completed_with_times:
            total_response_time = sum([
                (req.accepted_at - req.requested_at).total_seconds()
                for req in completed_with_times
            ])
            avg_response_time = total_response_time / len(completed_with_times)

        # Average completion time (time from REQUEST to complete - TOPLAM SÜRE)
        avg_completion_time = 0
        completed_with_completion_times = [req for req in completed_with_times if req.completed_at and req.requested_at]
        if completed_with_completion_times:
            total_completion_time = sum([
                (req.completed_at - req.requested_at).total_seconds()
                for req in completed_with_completion_times
            ])
            avg_completion_time = total_completion_time / len(completed_with_completion_times) if total_completion_time else 0

        return {
            'date': date.strftime('%Y-%m-%d'),
            'total_requests': total_requests,
            'completed_requests': completed_requests,
            'cancelled_requests': cancelled_requests,
            'PENDING_requests': PENDING_requests,
            'completion_rate': round((completed_requests / total_requests * 100) if total_requests > 0 else 0, 2),
            'avg_response_time_seconds': round(avg_response_time, 2),
            'avg_response_time_minutes': round(avg_response_time / 60, 2),
            'avg_completion_time_seconds': round(avg_completion_time, 2),
            'avg_completion_time_minutes': round(avg_completion_time / 60, 2)
        }

    @staticmethod
    def get_buggy_performance(hotel_id: int,
                             buggy_id: Optional[int] = None,
                             start_date: Optional[datetime] = None,
                             end_date: Optional[datetime] = None) -> List[Dict[str, Any]]:
        """
        Generate buggy performance report

        Args:
            hotel_id: Hotel ID
            buggy_id: Specific buggy ID (None for all buggies)
            start_date: Start date (defaults to 7 days ago)
            end_date: End date (defaults to now)

        Returns:
            List of buggy performance dictionaries
        """
        if start_date is None:
            start_date = datetime.utcnow() - timedelta(days=7)
        if end_date is None:
            end_date = datetime.utcnow()

        # Build query
        query = Buggy.query.filter(Buggy.hotel_id == hotel_id)
        if buggy_id:
            query = query.filter(Buggy.id == buggy_id)

        buggies = query.all()

        results = []
        for buggy in buggies:
            # Get requests for this buggy
            requests = BuggyRequest.query.filter(
                BuggyRequest.buggy_id == buggy.id,
                BuggyRequest.requested_at >= start_date,
                BuggyRequest.requested_at <= end_date
            ).all()

            completed = [r for r in requests if r.status == RequestStatus.COMPLETED]

            # Calculate metrics
            total_requests = len(requests)
            completed_requests = len(completed)

            avg_response = 0
            avg_completion = 0

            if completed:
                response_times = [
                    (r.accepted_at - r.requested_at).total_seconds()
                    for r in completed if r.accepted_at
                ]
                if response_times:
                    avg_response = sum(response_times) / len(response_times)

                completion_times = [
                    (r.completed_at - r.requested_at).total_seconds()
                    for r in completed if r.completed_at and r.requested_at
                ]
                if completion_times:
                    avg_completion = sum(completion_times) / len(completion_times)

            results.append({
                'buggy_id': buggy.id,
                'buggy_code': buggy.code,
                'driver_name': buggy.driver.full_name if buggy.driver else None,
                'total_requests': total_requests,
                'completed_requests': completed_requests,
                'completion_rate': round((completed_requests / total_requests * 100) if total_requests > 0 else 0, 2),
                'avg_response_time_minutes': round(avg_response / 60, 2),
                'avg_completion_time_minutes': round(avg_completion / 60, 2),
                'current_status': buggy.status.value if buggy.status else 'offline'
            })

        return results

    @staticmethod
    def get_location_analytics(hotel_id: int,
                               start_date: Optional[datetime] = None,
                               end_date: Optional[datetime] = None) -> List[Dict[str, Any]]:
        """
        Generate location analytics report

        Args:
            hotel_id: Hotel ID
            start_date: Start date (defaults to 30 days ago)
            end_date: End date (defaults to now)

        Returns:
            List of location analytics dictionaries
        """
        if start_date is None:
            start_date = datetime.utcnow() - timedelta(days=30)
        if end_date is None:
            end_date = datetime.utcnow()

        # Get all active locations
        locations = Location.query.filter_by(hotel_id=hotel_id, is_active=True).all()

        results = []
        for location in locations:
            # Get requests from this location
            requests = BuggyRequest.query.filter(
                BuggyRequest.location_id == location.id,
                BuggyRequest.requested_at >= start_date,
                BuggyRequest.requested_at <= end_date
            ).all()

            total_requests = len(requests)
            completed = len([r for r in requests if r.status == RequestStatus.COMPLETED])

            # Calculate average wait time
            completed_with_times = [r for r in requests if r.status == RequestStatus.COMPLETED and r.accepted_at]
            avg_wait = 0
            if completed_with_times:
                wait_times = [(r.accepted_at - r.requested_at).total_seconds() for r in completed_with_times]
                avg_wait = sum(wait_times) / len(wait_times)

            # Hourly distribution
            hourly_dist = {}
            for hour in range(24):
                hourly_dist[f"{hour:02d}:00"] = 0

            for req in requests:
                hour = req.requested_at.hour
                hourly_dist[f"{hour:02d}:00"] += 1

            results.append({
                'location_id': location.id,
                'location_name': location.name,
                'total_requests': total_requests,
                'completed_requests': completed,
                'completion_rate': round((completed / total_requests * 100) if total_requests > 0 else 0, 2),
                'avg_wait_time_minutes': round(avg_wait / 60, 2),
                'hourly_distribution': hourly_dist,
                'peak_hour': max(hourly_dist.items(), key=lambda x: x[1])[0] if hourly_dist else None
            })

        # Sort by total requests descending
        results.sort(key=lambda x: x['total_requests'], reverse=True)

        return results

    @staticmethod
    def get_request_details(hotel_id: int,
                           status: Optional[RequestStatus] = None,
                           start_date: Optional[datetime] = None,
                           end_date: Optional[datetime] = None,
                           limit: int = 100) -> List[Dict[str, Any]]:
        """
        Get detailed request list

        Args:
            hotel_id: Hotel ID
            status: Filter by status (None for all)
            start_date: Start date
            end_date: End date
            limit: Maximum number of records

        Returns:
            List of request details
        """
        query = BuggyRequest.query.filter(BuggyRequest.hotel_id == hotel_id)

        if status:
            query = query.filter(BuggyRequest.status == status)

        if start_date:
            query = query.filter(BuggyRequest.requested_at >= start_date)

        if end_date:
            query = query.filter(BuggyRequest.requested_at <= end_date)

        requests = query.order_by(BuggyRequest.requested_at.desc()).limit(limit).all()

        results = []
        for req in requests:
            results.append({
                'id': req.id,
                'location_name': req.location.name if req.location else None,
                'buggy_code': req.buggy.code if req.buggy else None,
                'driver_name': req.buggy.driver.full_name if req.buggy and req.buggy.driver else None,
                'room_number': req.room_number,
                'status': req.status.value,
                'requested_at': req.requested_at.isoformat() if req.requested_at else None,
                'accepted_at': req.accepted_at.isoformat() if req.accepted_at else None,
                'completed_at': req.completed_at.isoformat() if req.completed_at else None,
                'response_time_seconds': (req.accepted_at - req.requested_at).total_seconds() if req.accepted_at else None,
                'completion_time_seconds': (req.completed_at - req.accepted_at).total_seconds() if req.completed_at and req.accepted_at else None,
                'notes': req.notes
            })

        return results


    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str, sheet_name: str = 'Report') -> bytes:
        """
        Export data to Excel format
        
        Args:
            data: List of dictionaries to export
            filename: Name of the file
            sheet_name: Name of the sheet
            
        Returns:
            Excel file as bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        if not data:
            raise ValueError("No data to export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                # Convert datetime to string
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], title: str, filename: str) -> bytes:
        """
        Export data to PDF format with Turkish character support
        
        Args:
            data: List of dictionaries to export
            title: Report title
            filename: Name of the file
            
        Returns:
            PDF file as bytes
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import os
        
        if not data:
            raise ValueError("No data to export")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Türkçe karakter desteği için font kaydet
        # Projedeki DejaVu Sans fontlarını kullan
        try:
            # Proje içindeki font yolu
            from flask import current_app
            font_dir = os.path.join(current_app.root_path, 'static', 'fonts')
            
            font_regular = os.path.join(font_dir, 'DejaVuSans.ttf')
            font_bold = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
            
            if os.path.exists(font_regular) and os.path.exists(font_bold):
                pdfmetrics.registerFont(TTFont('DejaVuSans', font_regular))
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_bold))
                font_name = 'DejaVuSans'
                font_name_bold = 'DejaVuSans-Bold'
            else:
                # Fallback: Helvetica (sınırlı Türkçe desteği)
                font_name = 'Helvetica'
                font_name_bold = 'Helvetica-Bold'
        except Exception as e:
            # Fallback: Helvetica
            print(f"⚠️ Font yüklenemedi: {str(e)}")
            font_name = 'Helvetica'
            font_name_bold = 'Helvetica-Bold'
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=font_name_bold,
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data:
            row_data = []
            for header in headers:
                value = row.get(header, '')
                # Convert datetime to string
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                row_data.append(str(value))
            table_data.append(row_data)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()


    @staticmethod
    def get_advanced_analytics(hotel_id: int, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """
        Get advanced analytics with detailed breakdowns
        
        Args:
            hotel_id: Hotel ID
            start_date: Start date for analysis
            end_date: End date for analysis
            
        Returns:
            Comprehensive analytics dictionary
        """
        # Get all requests in date range
        all_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.requested_at >= start_date,
            BuggyRequest.requested_at < end_date
        ).all()
        
        # Initialize counters
        total_requests = len(all_requests)
        completed_count = 0
        cancelled_count = 0
        unanswered_count = 0
        PENDING_count = 0
        
        # Breakdown by cancellation reason
        cancelled_by_driver = 0
        cancelled_by_guest = 0
        cancelled_by_admin = 0
        
        # Response time analysis
        response_times = []
        completion_times = []
        
        # Location analysis
        location_stats = {}
        
        # Hour of day analysis
        hour_stats = {str(i): 0 for i in range(24)}
        
        # Day of week analysis
        day_stats = {
            'Monday': 0, 'Tuesday': 0, 'Wednesday': 0, 'Thursday': 0,
            'Friday': 0, 'Saturday': 0, 'Sunday': 0
        }
        
        # Process each request
        for req in all_requests:
            # Status counts
            if req.status == RequestStatus.COMPLETED:
                completed_count += 1
                
                # Response time
                if req.response_time:
                    response_times.append(req.response_time)
                
                # Completion time - Dinamik hesaplama (requested_at -> completed_at)
                completion_time = None
                if req.completed_at and req.requested_at:
                    delta = req.completed_at - req.requested_at
                    completion_time = int(delta.total_seconds())
                elif req.completion_time:
                    # Fallback: Veritabanındaki değeri kullan
                    completion_time = req.completion_time
                
                if completion_time and completion_time > 0:
                    completion_times.append(completion_time)
            elif req.status == RequestStatus.CANCELLED:
                cancelled_count += 1
                if req.cancelled_by == 'driver':
                    cancelled_by_driver += 1
                elif req.cancelled_by == 'guest':
                    cancelled_by_guest += 1
                elif req.cancelled_by == 'admin':
                    cancelled_by_admin += 1
            elif req.status == RequestStatus.UNANSWERED:
                unanswered_count += 1
            elif req.status == RequestStatus.PENDING:
                PENDING_count += 1
            
            # Location stats
            if req.location:
                loc_name = req.location.name
                if loc_name not in location_stats:
                    location_stats[loc_name] = {
                        'total': 0,
                        'completed': 0,
                        'cancelled': 0,
                        'unanswered': 0
                    }
                location_stats[loc_name]['total'] += 1
                if req.status == RequestStatus.COMPLETED:
                    location_stats[loc_name]['completed'] += 1
                elif req.status == RequestStatus.CANCELLED:
                    location_stats[loc_name]['cancelled'] += 1
                elif req.status == RequestStatus.UNANSWERED:
                    location_stats[loc_name]['unanswered'] += 1
            
            # Hour of day stats
            if req.requested_at:
                hour = str(req.requested_at.hour)
                hour_stats[hour] += 1
                
                # Day of week stats
                day_name = req.requested_at.strftime('%A')
                day_stats[day_name] += 1
        
        # Calculate averages
        avg_response_time = sum(response_times) / len(response_times) if response_times else 0
        avg_completion_time = sum(completion_times) / len(completion_times) if completion_times else 0
        
        # Calculate percentages
        completion_rate = (completed_count / total_requests * 100) if total_requests > 0 else 0
        cancellation_rate = (cancelled_count / total_requests * 100) if total_requests > 0 else 0
        unanswered_rate = (unanswered_count / total_requests * 100) if total_requests > 0 else 0
        
        # Find peak hour
        peak_hour = max(hour_stats.items(), key=lambda x: x[1])[0] if any(hour_stats.values()) else '0'
        
        # Find busiest day
        busiest_day = max(day_stats.items(), key=lambda x: x[1])[0] if any(day_stats.values()) else 'Monday'
        
        # Sort locations by total requests
        top_locations = sorted(
            location_stats.items(),
            key=lambda x: x[1]['total'],
            reverse=True
        )[:10]  # Top 10 locations
        
        return {
            'summary': {
                'total_requests': total_requests,
                'completed': completed_count,
                'cancelled': cancelled_count,
                'unanswered': unanswered_count,
                'PENDING': PENDING_count,
                'completion_rate': round(completion_rate, 2),
                'cancellation_rate': round(cancellation_rate, 2),
                'unanswered_rate': round(unanswered_rate, 2)
            },
            'cancellation_breakdown': {
                'by_driver': cancelled_by_driver,
                'by_guest': cancelled_by_guest,
                'by_admin': cancelled_by_admin
            },
            'performance': {
                'avg_response_time_seconds': round(avg_response_time, 2),
                'avg_response_time_minutes': round(avg_response_time / 60, 2),
                'avg_completion_time_seconds': round(avg_completion_time, 2),
                'avg_completion_time_minutes': round(avg_completion_time / 60, 2)
            },
            'peak_analysis': {
                'peak_hour': peak_hour,
                'busiest_day': busiest_day,
                'hourly_distribution': hour_stats,
                'daily_distribution': day_stats
            },
            'location_analysis': {
                'top_locations': [
                    {
                        'name': loc[0],
                        'total': loc[1]['total'],
                        'completed': loc[1]['completed'],
                        'cancelled': loc[1]['cancelled'],
                        'unanswered': loc[1]['unanswered'],
                        'completion_rate': round(loc[1]['completed'] / loc[1]['total'] * 100, 2) if loc[1]['total'] > 0 else 0
                    }
                    for loc in top_locations
                ]
            },
            'unanswered_details': {
                'total': unanswered_count,
                'percentage': round(unanswered_rate, 2),
                'note': 'Requests that were not answered within 1 hour'
            }
        }


    @staticmethod
    def get_route_analytics(hotel_id: int, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """
        Rota analizi - Başlangıç ve bitiş konumları arası istatistikler
        
        Args:
            hotel_id: Hotel ID
            start_date: Başlangıç tarihi
            end_date: Bitiş tarihi
            
        Returns:
            Rota istatistikleri
        """
        # Tamamlanmış ve completion_location_id olan talepleri al
        completed_requests = BuggyRequest.query.filter(
            BuggyRequest.hotel_id == hotel_id,
            BuggyRequest.status == RequestStatus.COMPLETED,
            BuggyRequest.completion_location_id.isnot(None),
            BuggyRequest.requested_at >= start_date,
            BuggyRequest.requested_at < end_date
        ).all()
        
        # Rota istatistikleri
        route_stats = {}
        driver_stats = {}
        buggy_stats = {}
        
        for req in completed_requests:
            # Rota key: "Başlangıç -> Bitiş"
            start_loc = req.location.name if req.location else "Bilinmiyor"
            end_loc = req.completion_location.name if req.completion_location else "Bilinmiyor"
            route_key = f"{start_loc} → {end_loc}"
            
            # Rota istatistiklerini güncelle
            if route_key not in route_stats:
                route_stats[route_key] = {
                    'start_location': start_loc,
                    'end_location': end_loc,
                    'start_location_id': req.location_id,
                    'end_location_id': req.completion_location_id,
                    'count': 0,
                    'total_time': 0,
                    'times': []
                }
            
            route_stats[route_key]['count'] += 1
            
            # Tamamlanma süresini hesapla (requested_at -> completed_at)
            # Veritabanındaki değer yerine dinamik hesaplama yap
            completion_time = None
            if req.completed_at and req.requested_at:
                delta = req.completed_at - req.requested_at
                completion_time = int(delta.total_seconds())
            elif req.completion_time:
                # Fallback: Veritabanındaki değeri kullan
                completion_time = req.completion_time
            
            if completion_time and completion_time > 0:
                route_stats[route_key]['total_time'] += completion_time
                route_stats[route_key]['times'].append(completion_time)
            
            # Sürücü istatistikleri
            if req.accepted_by_driver:
                driver_name = req.accepted_by_driver.full_name
                if driver_name not in driver_stats:
                    driver_stats[driver_name] = {
                        'driver_id': req.accepted_by_id,
                        'total_completed': 0,
                        'total_time': 0,
                        'routes': {}
                    }
                driver_stats[driver_name]['total_completed'] += 1
                if completion_time and completion_time > 0:
                    driver_stats[driver_name]['total_time'] += completion_time
                
                # Sürücünün rota dağılımı
                if route_key not in driver_stats[driver_name]['routes']:
                    driver_stats[driver_name]['routes'][route_key] = 0
                driver_stats[driver_name]['routes'][route_key] += 1
            
            # Buggy istatistikleri
            if req.buggy:
                buggy_code = req.buggy.code
                if buggy_code not in buggy_stats:
                    buggy_stats[buggy_code] = {
                        'buggy_id': req.buggy_id,
                        'total_completed': 0,
                        'total_time': 0
                    }
                buggy_stats[buggy_code]['total_completed'] += 1
                if completion_time and completion_time > 0:
                    buggy_stats[buggy_code]['total_time'] += completion_time
        
        # Ortalama süreleri hesapla ve sırala
        route_list = []
        for route_key, stats in route_stats.items():
            avg_time = stats['total_time'] / stats['count'] if stats['count'] > 0 else 0
            min_time = min(stats['times']) if stats['times'] else 0
            max_time = max(stats['times']) if stats['times'] else 0
            
            # Debug log
            #print(f"🛣️ Rota: {route_key}")
            #print(f"   Kullanım: {stats['count']} kez")
            #print(f"   Toplam süre: {stats['total_time']} saniye")
            #print(f"   Ortalama: {avg_time} saniye = {avg_time / 60} dakika")
            #print(f"   Süreler: {stats['times']}")
            
            route_list.append({
                'route': route_key,
                'start_location': stats['start_location'],
                'end_location': stats['end_location'],
                'start_location_id': stats['start_location_id'],
                'end_location_id': stats['end_location_id'],
                'count': stats['count'],
                'avg_time_seconds': round(avg_time, 2),
                'avg_time_minutes': round(avg_time / 60, 2),
                'min_time_seconds': round(min_time, 2),
                'max_time_seconds': round(max_time, 2)
            })
        
        # En çok kullanılan rotalara göre sırala
        route_list.sort(key=lambda x: x['count'], reverse=True)
        
        # Sürücü performansını hesapla
        driver_list = []
        for driver_name, stats in driver_stats.items():
            avg_time = stats['total_time'] / stats['total_completed'] if stats['total_completed'] > 0 else 0
            most_used_route = max(stats['routes'].items(), key=lambda x: x[1])[0] if stats['routes'] else None
            
            driver_list.append({
                'driver_name': driver_name,
                'driver_id': stats['driver_id'],
                'total_completed': stats['total_completed'],
                'avg_completion_time_minutes': round(avg_time / 60, 2),
                'most_used_route': most_used_route,
                'route_count': len(stats['routes'])
            })
        
        # Tamamlanan talep sayısına göre sırala
        driver_list.sort(key=lambda x: x['total_completed'], reverse=True)
        
        # Buggy performansını hesapla
        buggy_list = []
        for buggy_code, stats in buggy_stats.items():
            avg_time = stats['total_time'] / stats['total_completed'] if stats['total_completed'] > 0 else 0
            
            buggy_list.append({
                'buggy_code': buggy_code,
                'buggy_id': stats['buggy_id'],
                'total_completed': stats['total_completed'],
                'avg_completion_time_minutes': round(avg_time / 60, 2)
            })
        
        buggy_list.sort(key=lambda x: x['total_completed'], reverse=True)
        
        return {
            'total_completed_with_routes': len(completed_requests),
            'unique_routes': len(route_list),
            'most_popular_routes': route_list[:10],  # Top 10
            'all_routes': route_list,
            'driver_performance': driver_list,
            'buggy_performance': buggy_list,
            'date_range': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            }
        }
